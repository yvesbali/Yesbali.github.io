#!/usr/bin/env python3
"""
send_daily_email.py — Email quotidien de suivi LCDMH
=====================================================
Lit la dernière ligne de data/baselines/daily_stats_log.xlsx, calcule le delta
vs l'avant-dernière ligne, construit un email court et l'envoie via Gmail SMTP.

Variables d'environnement requises :
  GMAIL_USER         : yvesbali@gmail.com (défaut)
  GMAIL_APP_PASSWORD : mot de passe d'application Gmail (16 car.)
                       Génération : https://myaccount.google.com/apppasswords

Optionnel :
  EMAIL_RECIPIENT    : destinataire (défaut = GMAIL_USER)
  DAILY_XLSX_PATH    : chemin du xlsx (défaut data/baselines/daily_stats_log.xlsx)

Usage :
  python scripts/send_daily_email.py          # envoi réel
  python scripts/send_daily_email.py --dry    # affiche l'email sans envoyer
"""
import argparse
import os
import smtplib
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERREUR : openpyxl manquant. pip install openpyxl")
    sys.exit(1)

BASE = Path(__file__).resolve().parent.parent
XLSX = Path(os.getenv("DAILY_XLSX_PATH", str(BASE / "data/baselines/daily_stats_log.xlsx")))

GMAIL_USER = os.getenv("GMAIL_USER", "yvesbali@gmail.com")
GMAIL_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
RECIPIENT = os.getenv("EMAIL_RECIPIENT", GMAIL_USER)

HEADERS = [
    "Date", "videos_total", "videos_with_data", "sum_views_30d",
    "sum_impressions", "avg_ctr_weighted", "avg_view_pct_mean",
    "gsc_queries", "gsc_clicks", "gsc_impressions", "gsc_ctr_pct",
    "gsc_top3", "gsc_top10", "gsc_top20", "notes",
]


def read_last_two_rows() -> tuple:
    if not XLSX.exists():
        return None, None
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Log"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and any(v is not None for v in r):
            rows.append(dict(zip(HEADERS, r)))
    if not rows:
        return None, None
    last = rows[-1]
    prev = rows[-2] if len(rows) >= 2 else None
    return last, prev


def _fmt_date(v):
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v) if v is not None else "?"


def _num(v, default=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _delta(cur, prev, key):
    if prev is None or prev.get(key) is None or cur.get(key) is None:
        return None
    return _num(cur[key]) - _num(prev[key])


def _fmt_delta(d, unit="", digits=0):
    if d is None:
        return "—"
    sign = "+" if d >= 0 else ""
    if digits == 0:
        return f"{sign}{int(d)}{unit}"
    return f"{sign}{d:.{digits}f}{unit}"


def build_email(last: dict, prev: dict) -> tuple:
    date_s = _fmt_date(last.get("Date"))
    subject = f"LCDMH — Daily tracking {date_s}"

    # deltas
    dv = _delta(last, prev, "sum_views_30d")
    di = _delta(last, prev, "sum_impressions")
    dctr = _delta(last, prev, "avg_ctr_weighted")
    dclicks = _delta(last, prev, "gsc_clicks")
    dimpr_gsc = _delta(last, prev, "gsc_impressions")
    dtop10 = _delta(last, prev, "gsc_top10")
    dtop3 = _delta(last, prev, "gsc_top3")

    prev_label = f"vs {_fmt_date(prev.get('Date'))}" if prev else "(première ligne — pas de Δ)"

    # TXT
    txt = f"""LCDMH — Daily tracking {date_s}
{'=' * 44}

YOUTUBE (snapshot seo_stats.json)
---------------------------------
Vidéos suivies     : {last.get('videos_total')}  (dont {last.get('videos_with_data')} avec données)
Vues cumulées 30j  : {int(_num(last.get('sum_views_30d'))):,}   Δ {_fmt_delta(dv, ' vues')}
Impressions        : {int(_num(last.get('sum_impressions'))):,}   Δ {_fmt_delta(di, ' impr')}
CTR moyen pondéré  : {_num(last.get('avg_ctr_weighted'))} %   Δ {_fmt_delta(dctr, ' pts', 2)}
Rétention moyenne  : {_num(last.get('avg_view_pct_mean'))} %

GOOGLE SEARCH CONSOLE (dernier export)
--------------------------------------
Requêtes           : {last.get('gsc_queries')}
Clics              : {last.get('gsc_clicks')}   Δ {_fmt_delta(dclicks, ' clics')}
Impressions        : {last.get('gsc_impressions')}   Δ {_fmt_delta(dimpr_gsc, ' impr')}
CTR                : {last.get('gsc_ctr_pct')} %
Top 3 / 10 / 20    : {last.get('gsc_top3')} / {last.get('gsc_top10')} / {last.get('gsc_top20')}
                     Δ top10 {_fmt_delta(dtop10, ' reqs')}   Δ top3 {_fmt_delta(dtop3, ' reqs')}

Source GSC : {last.get('notes')}
Référence  : {prev_label}

Dashboard complet : F:\\LCDMH_GitHub_Audit\\data\\baselines\\daily_stats_log.xlsx

--
LCDMH — tracking automatique
"""

    # HTML (simple, lisible sur mobile)
    def row(label, val, delta_txt=""):
        delta_html = ""
        if delta_txt and delta_txt != "—":
            color = "#27ae60" if delta_txt.startswith("+") and not delta_txt.startswith("+0") else (
                "#e74c3c" if delta_txt.startswith("-") else "#666")
            delta_html = f'<span style="color:{color};font-size:12px;margin-left:8px">Δ {delta_txt}</span>'
        return f'<tr><td style="padding:6px 12px;color:#555">{label}</td><td style="padding:6px 12px;font-weight:600">{val}{delta_html}</td></tr>'

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:-apple-system,Arial,sans-serif;max-width:620px;margin:0 auto;padding:16px;background:#f7f5f0;color:#222">

<div style="background:#1a2b4a;color:#fff;padding:18px;border-radius:10px 10px 0 0">
  <div style="font-size:13px;opacity:.85">Daily tracking</div>
  <div style="font-size:22px;font-weight:bold">🏍️ LCDMH — {date_s}</div>
</div>

<div style="background:#fff;padding:20px;border-radius:0 0 10px 10px;box-shadow:0 1px 4px rgba(0,0,0,.06)">

<h3 style="margin:0 0 8px;color:#1a2b4a">YouTube (snapshot)</h3>
<table style="width:100%;border-collapse:collapse;font-size:14px">
{row("Vidéos suivies", f"{last.get('videos_total')} (dont {last.get('videos_with_data')} avec data)")}
{row("Vues cumulées 30j", f"{int(_num(last.get('sum_views_30d'))):,}".replace(",", " "), _fmt_delta(dv, ""))}
{row("Impressions", f"{int(_num(last.get('sum_impressions'))):,}".replace(",", " "), _fmt_delta(di, ""))}
{row("CTR moyen pondéré", f"{_num(last.get('avg_ctr_weighted'))} %", _fmt_delta(dctr, " pts", 2))}
{row("Rétention moyenne", f"{_num(last.get('avg_view_pct_mean'))} %")}
</table>

<h3 style="margin:18px 0 8px;color:#1a2b4a">Google Search Console</h3>
<table style="width:100%;border-collapse:collapse;font-size:14px">
{row("Requêtes impressions", last.get('gsc_queries'))}
{row("Clics", last.get('gsc_clicks'), _fmt_delta(dclicks, ""))}
{row("Impressions", last.get('gsc_impressions'), _fmt_delta(dimpr_gsc, ""))}
{row("CTR", f"{last.get('gsc_ctr_pct')} %")}
{row("Top 3", last.get('gsc_top3'), _fmt_delta(dtop3, ""))}
{row("Top 10", last.get('gsc_top10'), _fmt_delta(dtop10, ""))}
{row("Top 20", last.get('gsc_top20'))}
</table>

<p style="font-size:12px;color:#888;margin:18px 0 0">
Référence Δ : {prev_label}<br>
Source GSC : {last.get('notes')}<br>
Dashboard complet : <code>data/baselines/daily_stats_log.xlsx</code>
</p>

</div>

<p style="text-align:center;font-size:11px;color:#aaa;margin-top:14px">
LCDMH — daily_stats automatique · <a href="https://lcdmh.com" style="color:#e67e22">lcdmh.com</a>
</p>

</body></html>"""

    return subject, txt, html


def send_email(subject: str, txt: str, html: str) -> bool:
    if not GMAIL_PASSWORD:
        print("[ERR] GMAIL_APP_PASSWORD non défini — email non envoyé.")
        print("      Crée un app password sur https://myaccount.google.com/apppasswords")
        print("      puis PowerShell :")
        print('        [Environment]::SetEnvironmentVariable("GMAIL_APP_PASSWORD", "xxxxxxxxxxxxxxxx", "User")')
        return False
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = GMAIL_USER
    msg["To"] = RECIPIENT
    msg.attach(MIMEText(txt, "plain", "utf-8"))
    msg.attach(MIMEText(html, "html", "utf-8"))
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_PASSWORD)
            s.sendmail(GMAIL_USER, RECIPIENT, msg.as_string())
        print(f"[OK] email envoyé -> {RECIPIENT}")
        return True
    except Exception as e:
        print(f"[ERR] envoi échoué : {e}")
        return False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry", action="store_true", help="affiche l'email sans l'envoyer")
    args = ap.parse_args()

    last, prev = read_last_two_rows()
    if last is None:
        print(f"[ERR] pas de ligne dans {XLSX} — lance d'abord append_daily_log.py")
        sys.exit(2)

    subject, txt, html = build_email(last, prev)

    if args.dry:
        print("=" * 60)
        print("DRY-RUN — email qui serait envoyé :")
        print("=" * 60)
        print(f"De     : {GMAIL_USER}")
        print(f"À      : {RECIPIENT}")
        print(f"Sujet  : {subject}")
        print()
        print(txt)
        return

    print(f"[INFO] préparation email daily — {datetime.now().strftime('%H:%M:%S')}")
    print(f"       destinataire : {RECIPIENT}")
    print(f"       sujet        : {subject}")
    send_email(subject, txt, html)


if __name__ == "__main__":
    main()
