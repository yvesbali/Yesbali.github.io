#!/usr/bin/env python3
"""
append_daily_log.py — Ajoute une ligne quotidienne dans daily_stats_log.xlsx
============================================================================
Lecture : seo_stats.json + dernier GSC CSV disponible
Écriture : data/baselines/daily_stats_log.xlsx (feuille "Log"), append en bas.

Évite les doublons : si la date du jour est déjà présente, la ligne est mise à jour.
"""

import csv, json, sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERREUR : openpyxl manquant. pip install openpyxl")
    sys.exit(1)

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / "data/baselines/daily_stats_log.xlsx"

HEADERS = [
    "Date", "videos_total", "videos_with_data", "sum_views_30d",
    "sum_impressions", "avg_ctr_weighted", "avg_view_pct_mean",
    "gsc_queries", "gsc_clicks", "gsc_impressions", "gsc_ctr_pct",
    "gsc_top3", "gsc_top10", "gsc_top20", "notes",
]


def read_seo():
    p = BASE / "seo_stats.json"
    if not p.exists():
        return None
    try:
        d = json.load(open(p, encoding="utf-8"))
    except Exception as e:
        print(f"Impossible de lire seo_stats.json : {e}")
        return None
    vids = []
    for vid, data in d.items():
        snaps = data.get("snapshots", [])
        best = None
        for s in snaps:
            if s.get("views_30d", 0) > 0 or s.get("impressions", 0) > 0:
                if best is None or s.get("taken_at", "") > best.get("taken_at", ""):
                    best = s
        if best is None and snaps:
            best = snaps[-1]
        vids.append({
            "video_id": vid,
            "views_30d": (best or {}).get("views_30d", 0),
            "impressions": (best or {}).get("impressions", 0),
            "ctr_pct": (best or {}).get("ctr", 0),
            "avg_view_pct": (best or {}).get("avg_view_pct", 0),
        })
    non_null = [v for v in vids if v["views_30d"] > 0 or v["impressions"] > 0]
    sum_views = sum(v["views_30d"] for v in vids)
    sum_impr = sum(v["impressions"] for v in vids)
    ctr_w = round(sum(v["ctr_pct"] * v["impressions"] for v in vids) / sum_impr, 2) if sum_impr else 0
    avg_vp = round(sum(v["avg_view_pct"] for v in non_null) / len(non_null), 1) if non_null else 0
    return {
        "videos_total": len(vids),
        "videos_with_data": len(non_null),
        "sum_views_30d": sum_views,
        "sum_impressions": sum_impr,
        "avg_ctr_weighted": ctr_w,
        "avg_view_pct_mean": avg_vp,
    }


def read_latest_gsc():
    bdir = BASE / "data/baselines"
    candidates = sorted(bdir.glob("gsc_queries_*.csv"))
    if not candidates:
        return None
    p = candidates[-1]
    try:
        rows = list(csv.DictReader(open(p, encoding="utf-8")))
    except Exception:
        return None
    tot_clicks = sum(int(r.get("clicks", 0) or 0) for r in rows)
    tot_impr = sum(int(r.get("impressions", 0) or 0) for r in rows)
    ctr = round(tot_clicks / tot_impr * 100, 2) if tot_impr else 0
    top3 = top10 = top20 = 0
    for r in rows:
        try:
            pos = float(r.get("position", 0) or 0)
        except Exception:
            continue
        if 0 < pos <= 3:
            top3 += 1
        if 0 < pos <= 10:
            top10 += 1
        if 0 < pos <= 20:
            top20 += 1
    return {
        "gsc_file": p.name,
        "gsc_queries": len(rows),
        "gsc_clicks": tot_clicks,
        "gsc_impressions": tot_impr,
        "gsc_ctr_pct": ctr,
        "gsc_top3": top3,
        "gsc_top10": top10,
        "gsc_top20": top20,
    }


def ensure_workbook(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 30
    widths = [12, 12, 14, 14, 14, 14, 16, 12, 12, 14, 12, 10, 12, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(row=1, column=1).number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return load_workbook(path)


def main():
    yt = read_seo() or {}
    gsc = read_latest_gsc() or {}
    today = date.today().isoformat()
    row = [
        today,
        yt.get("videos_total"), yt.get("videos_with_data"),
        yt.get("sum_views_30d"), yt.get("sum_impressions"),
        yt.get("avg_ctr_weighted"), yt.get("avg_view_pct_mean"),
        gsc.get("gsc_queries"), gsc.get("gsc_clicks"),
        gsc.get("gsc_impressions"), gsc.get("gsc_ctr_pct"),
        gsc.get("gsc_top3"), gsc.get("gsc_top10"), gsc.get("gsc_top20"),
        f"gsc={gsc.get('gsc_file', '-')}",
    ]

    wb = ensure_workbook(XLSX)
    ws = wb["Log"]

    # Trouve la dernière ligne contenant une vraie valeur, et détecte si today est déjà présent.
    last_real_row = 1
    existing_row = None
    for r in range(2, ws.max_row + 1):
        vals_in_row = [ws.cell(row=r, column=c).value for c in range(1, len(HEADERS) + 1)]
        if any(v is not None for v in vals_in_row):
            last_real_row = r
            val = ws.cell(row=r, column=1).value
            v = val.isoformat() if hasattr(val, "isoformat") else (str(val) if val is not None else None)
            if v == today:
                existing_row = r
    target_row = existing_row if existing_row is not None else last_real_row + 1

    for i, val in enumerate(row, start=1):
        ws.cell(row=target_row, column=i, value=val)
    ws.cell(row=target_row, column=1).number_format = "yyyy-mm-dd"
    wb.save(XLSX)
    print(f"OK ligne {target_row} du {today} ecrite dans {XLSX}")


if __name__ == "__main__":
    main()
